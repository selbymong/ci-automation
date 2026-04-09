"""F083: Import Summer 2024 and Priority Reference historical data.

Imports prior cycle data for trend analysis, maps historical charities to
canonical records, imports SRSS year columns as SRSSHistorical, and imports
Priority Reference priority calculations.

Usage:
    python scripts/import/import_historical.py [--excel PATH]

Requires: openpyxl
"""
import argparse
import asyncio
import sys
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "backend"))

import openpyxl  # noqa: E402
from sqlalchemy import select  # noqa: E402
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine  # noqa: E402

from app.config import settings  # noqa: E402
from app.models.charity import Charity  # noqa: E402
from app.models.cycle import EvaluationCycle  # noqa: E402
from app.models.evaluation import Evaluation, EvaluationStage  # noqa: E402
from app.models.priority import PriorityScore  # noqa: E402
from app.models.srss import SRSSHistorical, percentage_to_grade  # noqa: E402

DEFAULT_EXCEL = Path(__file__).resolve().parents[1] / "data" / "Charities_in_Process_2025.xlsx"

# SRSS year columns: "SRSS 2013", "SRSS 2014", ..., "SRSS 2024"
SRSS_YEAR_RANGE = range(2013, 2025)

# Category pct columns per year (may or may not be present)
SRSS_CATEGORY_KEYS = [
    "Strategy %", "Activities %", "Outputs %",
    "Outcomes %", "Quality %", "Learning %",
]


def safe_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def safe_float(val) -> Optional[float]:
    if val is None or val == "" or val == "N/A":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_int(val) -> Optional[int]:
    if val is None or val == "" or val == "N/A":
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


async def import_summer_2024(
    records: list[dict],
    db: AsyncSession,
    cycle_name: str = "Summer 2024",
) -> dict:
    """Import Summer 2024 charities as a prior cycle."""
    stats = {
        "evaluations_created": 0, "charities_linked": 0,
        "skipped": 0, "errors": 0,
    }

    # Get or create cycle
    result = await db.execute(
        select(EvaluationCycle).where(EvaluationCycle.name == cycle_name)
    )
    cycle = result.scalar_one_or_none()
    if not cycle:
        cycle = EvaluationCycle(name=cycle_name, status="closed")
        db.add(cycle)
        await db.flush()

    for rec in records:
        cra = safe_str(rec.get("BN/Registration Number"))
        if not cra:
            stats["skipped"] += 1
            continue

        try:
            result = await db.execute(
                select(Charity).where(Charity.cra_number == cra)
            )
            charity = result.scalar_one_or_none()
            if not charity:
                # Create charity stub if not in Name Reference
                formal_name = safe_str(rec.get("Formal Name")) or f"Unknown ({cra})"
                charity = Charity(cra_number=cra, formal_name=formal_name)
                db.add(charity)
                await db.flush()

            stats["charities_linked"] += 1

            # Check for duplicate evaluation
            result = await db.execute(
                select(Evaluation).where(
                    Evaluation.charity_id == charity.id,
                    Evaluation.cycle_id == cycle.id,
                )
            )
            if result.scalar_one_or_none():
                stats["skipped"] += 1
                continue

            stage = safe_str(rec.get("Status"))
            if stage and stage.lower() in ("done", "complete", "published"):
                stage_val = EvaluationStage.published.value
            else:
                stage_val = EvaluationStage.published.value  # Assume completed for prior cycle

            evaluation = Evaluation(
                charity_id=charity.id,
                cycle_id=cycle.id,
                stage=stage_val,
            )
            db.add(evaluation)
            stats["evaluations_created"] += 1

        except Exception as e:
            print(f"  Error importing Summer 2024 {cra}: {e}")
            await db.rollback()
            stats["errors"] += 1

    await db.flush()
    return stats


async def import_srss_historical(
    records: list[dict],
    db: AsyncSession,
) -> dict:
    """Import SRSS historical scores from year columns."""
    stats = {"records_created": 0, "skipped": 0, "errors": 0}

    for rec in records:
        cra = safe_str(rec.get("BN/Registration Number"))
        if not cra:
            stats["skipped"] += 1
            continue

        try:
            result = await db.execute(
                select(Charity).where(Charity.cra_number == cra)
            )
            charity = result.scalar_one_or_none()
            if not charity:
                stats["skipped"] += 1
                continue

            for year in SRSS_YEAR_RANGE:
                # Try column formats: "SRSS 2013", "2013 SRSS", "SRSS_2013"
                total_pct = None
                for fmt in [f"SRSS {year}", f"{year} SRSS", f"SRSS_{year}", str(year)]:
                    total_pct = safe_float(rec.get(fmt))
                    if total_pct is not None:
                        break

                if total_pct is None:
                    continue

                # Check for duplicate
                result = await db.execute(
                    select(SRSSHistorical).where(
                        SRSSHistorical.charity_id == charity.id,
                        SRSSHistorical.year == year,
                    )
                )
                if result.scalar_one_or_none():
                    continue

                grade = percentage_to_grade(total_pct)

                # Look for category breakdowns
                cat_data = {}
                for cat_key in SRSS_CATEGORY_KEYS:
                    year_cat_key = f"{year} {cat_key}"
                    val = safe_float(rec.get(year_cat_key))
                    if val is not None:
                        field = cat_key.lower().replace(" %", "_pct").replace(" ", "_")
                        cat_data[field] = val

                hist = SRSSHistorical(
                    charity_id=charity.id,
                    year=year,
                    total_pct=total_pct,
                    letter_grade=grade,
                    **cat_data,
                )
                db.add(hist)
                stats["records_created"] += 1

        except Exception as e:
            print(f"  Error importing SRSS history for {cra}: {e}")
            await db.rollback()
            stats["errors"] += 1

    await db.flush()
    return stats


async def import_priority_reference(
    records: list[dict],
    db: AsyncSession,
    cycle_name: str = "Priority Reference",
) -> dict:
    """Import Priority Reference calculations."""
    stats = {"priorities_created": 0, "skipped": 0, "errors": 0}

    # Get or create cycle
    result = await db.execute(
        select(EvaluationCycle).where(EvaluationCycle.name == cycle_name)
    )
    cycle = result.scalar_one_or_none()
    if not cycle:
        cycle = EvaluationCycle(name=cycle_name)
        db.add(cycle)
        await db.flush()

    for rec in records:
        cra = safe_str(rec.get("BN/Registration Number"))
        if not cra:
            stats["skipped"] += 1
            continue

        try:
            result = await db.execute(
                select(Charity).where(Charity.cra_number == cra)
            )
            charity = result.scalar_one_or_none()
            if not charity:
                stats["skipped"] += 1
                continue

            composite = safe_float(rec.get("Composite Score")) or safe_float(rec.get("Priority Score"))
            rank = safe_int(rec.get("Priority")) or safe_int(rec.get("Rank"))

            if composite is None and rank is None:
                stats["skipped"] += 1
                continue

            views = safe_float(rec.get("Views Score"))
            staleness = safe_float(rec.get("Staleness Score"))
            demand = safe_float(rec.get("Demand Score"))
            top100 = safe_float(rec.get("Top 100 Bonus"))

            priority = PriorityScore(
                charity_id=charity.id,
                views_score=views,
                staleness_score=staleness,
                demand_score=demand,
                top100_bonus=top100,
                composite_score=composite or 0.0,
                priority_rank=rank or 5,
            )
            db.add(priority)
            stats["priorities_created"] += 1

        except Exception as e:
            print(f"  Error importing priority for {cra}: {e}")
            await db.rollback()
            stats["errors"] += 1

    await db.flush()
    return stats


async def import_from_records(records: list[dict], db: AsyncSession) -> dict:
    """Import all historical data from pre-parsed records."""
    s2024 = await import_summer_2024(records, db)
    srss = await import_srss_historical(records, db)
    priority = await import_priority_reference(records, db)

    return {
        "summer_2024": s2024,
        "srss_historical": srss,
        "priority_reference": priority,
    }


async def import_from_excel(excel_path: Path) -> dict:
    """Import historical data from multiple Excel sheets."""
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

    engine = create_async_engine(settings.DATABASE_URL)
    session_factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    all_stats = {}

    async with session_factory() as session:
        # Import Summer 2024 sheet
        if "Summer 2024" in wb.sheetnames:
            ws = wb["Summer 2024"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h).strip() if h else "" for h in rows[0]]
                records = [{headers[i]: cell for i, cell in enumerate(row) if i < len(headers)} for row in rows[1:]]
                all_stats["summer_2024"] = await import_summer_2024(records, session)

        # Import SRSS historical from any sheet that has year columns
        for sheet_name in ["Name Reference", "JD Charities List", "Summer 2025"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    headers = [str(h).strip() if h else "" for h in rows[0]]
                    records = [{headers[i]: cell for i, cell in enumerate(row) if i < len(headers)} for row in rows[1:]]
                    all_stats[f"srss_{sheet_name}"] = await import_srss_historical(records, session)

        # Import Priority Reference sheet
        if "Priority Reference" in wb.sheetnames:
            ws = wb["Priority Reference"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h).strip() if h else "" for h in rows[0]]
                records = [{headers[i]: cell for i, cell in enumerate(row) if i < len(headers)} for row in rows[1:]]
                all_stats["priority_reference"] = await import_priority_reference(records, session)

        await session.commit()

    await engine.dispose()
    wb.close()
    return all_stats


async def main():
    parser = argparse.ArgumentParser(description="Import historical data")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"Excel file not found: {args.excel}")
        sys.exit(1)

    print(f"Importing from: {args.excel}")
    stats = await import_from_excel(args.excel)
    print("Import complete:")
    for section, section_stats in stats.items():
        print(f"\n  {section}:")
        for k, v in section_stats.items():
            print(f"    {k}: {v}")


if __name__ == "__main__":
    asyncio.run(main())
